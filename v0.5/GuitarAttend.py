import tkinter as tk
import tkinter.messagebox as messagebox
from openpyxl.styles import Alignment, PatternFill
import openpyxl


class Mainwindow():
    def __init__(self):
        # 誰的出席紀錄
        self.WhosAttend = dict()
        # 儲存表格的狀態(顏色)
        self.the_seat = dict()
        # 儲存表格顯示名子
        self.formname = dict()
        # 初始化表格顯示名子
        for j in range(5):
            day = ""
            if j == 0:
                day = "M"
            elif j == 1:
                day = "T"
            elif j == 2:
                day = "W"
            elif j == 3:
                day = "R"
            elif j == 4:
                day = "F"
            for i in range(4, 15):
                if i == 11:
                    self.formname[day + "A"] = ""
                elif i == 12:
                    self.formname[day + "B"] = ""
                elif i == 13:
                    self.formname[day + "C"] = ""
                elif i == 14:
                    self.formname[day + "D"] = ""
                else:
                    self.formname[day + str(i)] = ""
        # 控制表格顯示名子元件
        self.formtext = dict()

        # 儲存排了多少節
        self.howmuch = dict()

        self.addname = ""
        self.canHeight = 785
        self.canWidth = 1250
        self.window = tk.Tk()
        self.window.title("吉他社排班程式 made by Zhuming")
        self.window.geometry("1250x785")

        # self.backFrame = tk.Frame(self.window)
        # self.backFrame.pack(fill = "both", expand = True)
        # self.backFrame.propagate(0)
        # self.window.minsize(width=600, height=400)
        # self.window.maxsize(width=1280, height=960)
        # self.window.resizable(width=False, height=False)

        self.mycanvas = tk.Canvas(self.window, width=self.canWidth, height=self.canHeight)  # 創造一個Canvas 1250x785
        self.mycanvas.pack(fill=tk.BOTH, expand=tk.YES)  # 排列

        self.mycanvas.create_rectangle(10, 10, 670, 780)  # 畫外框

        self.labeltest = tk.Label(self.window, text="在下方打上想排的節數 大小寫皆可\n多個可用逗號分開(Ex. M5,r7,Fa)\n前面加上-代表要刪除的節數:", width=50)
        self.labeltest.place(x=952, y=437, anchor="center")

        self.list1 = tk.Listbox(self.mycanvas, height=20, width=35, selectmode="single")
        self.list1.place(x=832, y=30)
        self.list1.bind("<<ListboxSelect>>", self.listbox_service)

        # 以下為畫橫線 第一橫條 為第一條以此類推 每條間隔65
        self.mycanvas.create_line(10, 65, 670, 65)
        self.mycanvas.create_line(10, 130, 670, 130)
        self.mycanvas.create_line(10, 195, 670, 195)
        self.mycanvas.create_line(10, 260, 670, 260)
        self.mycanvas.create_line(10, 325, 670, 325)
        self.mycanvas.create_line(10, 390, 670, 390)
        self.mycanvas.create_line(10, 455, 670, 455)
        self.mycanvas.create_line(10, 520, 670, 520)
        self.mycanvas.create_line(10, 585, 670, 585)
        self.mycanvas.create_line(10, 650, 670, 650)
        self.mycanvas.create_line(10, 715, 670, 715)

        # 以下為畫直線 第一直條 為第一條以此類推 每條間隔110
        self.mycanvas.create_line(120, 10, 120, 780)
        self.mycanvas.create_line(230, 10, 230, 780)
        self.mycanvas.create_line(340, 10, 340, 780)
        self.mycanvas.create_line(450, 10, 450, 780)
        self.mycanvas.create_line(560, 10, 560, 780)

        # 以下為畫第一格的斜線
        self.mycanvas.create_line(10, 10, 120, 65)

        # 以下為橫行畫字 (固定)
        self.mycanvas.create_text(175, 38, text='Monday\nM', font=('Arial', 11), justify="center")
        self.mycanvas.create_text(285, 38, text='Tuesday\nT', font=('Arial', 11), justify="center")
        self.mycanvas.create_text(395, 38, text='Wednesday\nW', font=('Arial', 11), justify="center")
        self.mycanvas.create_text(505, 38, text='Thursday\nR', font=('Arial', 11), justify="center")
        self.mycanvas.create_text(615, 38, text='Friday\nF', font=('Arial', 11), justify="center")

        # 以下為直行畫字 (固定)
        self.mycanvas.create_text(40, 54, text='時間', font=('Arial', 11))
        self.mycanvas.create_text(95, 27, text='星期', font=('Arial', 11))
        self.mycanvas.create_text(65, 98, text='11:20 ~ 12:10\n第4節', font=('Arial', 11), justify="center")
        self.mycanvas.create_text(65, 163, text='12:10 ~ 13:10\n第5節', font=('Arial', 11), justify="center")
        self.mycanvas.create_text(65, 228, text='13:20 ~ 14:10\n第6節', font=('Arial', 11), justify="center")
        self.mycanvas.create_text(65, 293, text='14:20 ~ 15:10\n第7節', font=('Arial', 11), justify="center")
        self.mycanvas.create_text(65, 358, text='15:30 ~ 16:20\n第8節', font=('Arial', 11), justify="center")
        self.mycanvas.create_text(65, 423, text='16:30 ~ 17:20\n第9節', font=('Arial', 11), justify="center")
        self.mycanvas.create_text(65, 488, text='17:30 ~ 18:20\n第10節', font=('Arial', 11), justify="center")
        self.mycanvas.create_text(65, 553, text='18:25 ~ 19:15\n第A節', font=('Arial', 11), justify="center")
        self.mycanvas.create_text(65, 618, text='19:20 ~ 20:10\n第B節', font=('Arial', 11), justify="center")
        self.mycanvas.create_text(65, 683, text='20:15 ~ 21:05\n第C節', font=('Arial', 11), justify="center")
        self.mycanvas.create_text(65, 748, text='21:10 ~ 22:00\n第D節', font=('Arial', 11), justify="center")

        # 以下為畫字(非固定)
        self.M4text = self.mycanvas.create_text(175, 93, font=('Arial', 10), justify="center")
        self.M5text = self.mycanvas.create_text(175, 158, font=('Arial', 10), justify="center")
        self.M6text = self.mycanvas.create_text(175, 223, font=('Arial', 10), justify="center")
        self.M7text = self.mycanvas.create_text(175, 288, font=('Arial', 10), justify="center")
        self.M8text = self.mycanvas.create_text(175, 353, font=('Arial', 10), justify="center")
        self.M9text = self.mycanvas.create_text(175, 418, font=('Arial', 10), justify="center")
        self.M10text = self.mycanvas.create_text(175, 483, font=('Arial', 10), justify="center")
        self.MAtext = self.mycanvas.create_text(175, 548, font=('Arial', 10), justify="center")
        self.MBtext = self.mycanvas.create_text(175, 613, font=('Arial', 10), justify="center")
        self.MCtext = self.mycanvas.create_text(175, 678, font=('Arial', 10), justify="center")
        self.MDtext = self.mycanvas.create_text(175, 743, font=('Arial', 10), justify="center")
        self.T4text = self.mycanvas.create_text(285, 93, font=('Arial', 10), justify="center")
        self.T5text = self.mycanvas.create_text(285, 158, font=('Arial', 10), justify="center")
        self.T6text = self.mycanvas.create_text(285, 223, font=('Arial', 10), justify="center")
        self.T7text = self.mycanvas.create_text(285, 288, font=('Arial', 10), justify="center")
        self.T8text = self.mycanvas.create_text(285, 353, font=('Arial', 10), justify="center")
        self.T9text = self.mycanvas.create_text(285, 418, font=('Arial', 10), justify="center")
        self.T10text = self.mycanvas.create_text(285, 483, font=('Arial', 10), justify="center")
        self.TAtext = self.mycanvas.create_text(285, 548, font=('Arial', 10), justify="center")
        self.TBtext = self.mycanvas.create_text(285, 613, font=('Arial', 10), justify="center")
        self.TCtext = self.mycanvas.create_text(285, 678, font=('Arial', 10), justify="center")
        self.TDtext = self.mycanvas.create_text(285, 743, font=('Arial', 10), justify="center")
        self.W4text = self.mycanvas.create_text(395, 93, font=('Arial', 10), justify="center")
        self.W5text = self.mycanvas.create_text(395, 158, font=('Arial', 10), justify="center")
        self.W6text = self.mycanvas.create_text(395, 223, font=('Arial', 10), justify="center")
        self.W7text = self.mycanvas.create_text(395, 288, font=('Arial', 10), justify="center")
        self.W8text = self.mycanvas.create_text(395, 353, font=('Arial', 10), justify="center")
        self.W9text = self.mycanvas.create_text(395, 418, font=('Arial', 10), justify="center")
        self.W10text = self.mycanvas.create_text(395, 483, font=('Arial', 10), justify="center")
        self.WAtext = self.mycanvas.create_text(395, 548, font=('Arial', 10), justify="center")
        self.WBtext = self.mycanvas.create_text(395, 613, font=('Arial', 10), justify="center")
        self.WCtext = self.mycanvas.create_text(395, 678, font=('Arial', 10), justify="center")
        self.WDtext = self.mycanvas.create_text(395, 743, font=('Arial', 10), justify="center")
        self.R4text = self.mycanvas.create_text(505, 93, font=('Arial', 10), justify="center")
        self.R5text = self.mycanvas.create_text(505, 158, font=('Arial', 10), justify="center")
        self.R6text = self.mycanvas.create_text(505, 223, font=('Arial', 10), justify="center")
        self.R7text = self.mycanvas.create_text(505, 288, font=('Arial', 10), justify="center")
        self.R8text = self.mycanvas.create_text(505, 353, font=('Arial', 10), justify="center")
        self.R9text = self.mycanvas.create_text(505, 418, font=('Arial', 10), justify="center")
        self.R10text = self.mycanvas.create_text(505, 483, font=('Arial', 10), justify="center")
        self.RAtext = self.mycanvas.create_text(505, 548, font=('Arial', 10), justify="center")
        self.RBtext = self.mycanvas.create_text(505, 613, font=('Arial', 10), justify="center")
        self.RCtext = self.mycanvas.create_text(505, 678, font=('Arial', 10), justify="center")
        self.RDtext = self.mycanvas.create_text(505, 743, font=('Arial', 10), justify="center")
        self.F4text = self.mycanvas.create_text(615, 93, font=('Arial', 10), justify="center")
        self.F5text = self.mycanvas.create_text(615, 158, font=('Arial', 10), justify="center")
        self.F6text = self.mycanvas.create_text(615, 223, font=('Arial', 10), justify="center")
        self.F7text = self.mycanvas.create_text(615, 288, font=('Arial', 10), justify="center")
        self.F8text = self.mycanvas.create_text(615, 353, font=('Arial', 10), justify="center")
        self.F9text = self.mycanvas.create_text(615, 418, font=('Arial', 10), justify="center")
        self.F10text = self.mycanvas.create_text(615, 483, font=('Arial', 10), justify="center")
        self.FAtext = self.mycanvas.create_text(615, 548, font=('Arial', 10), justify="center")
        self.FBtext = self.mycanvas.create_text(615, 613, font=('Arial', 10), justify="center")
        self.FCtext = self.mycanvas.create_text(615, 678, font=('Arial', 10), justify="center")
        self.FDtext = self.mycanvas.create_text(615, 743, font=('Arial', 10), justify="center")
        self.textnum = self.mycanvas.create_text(956, 390, fill="magenta", font=("Arial", 15), text="此人目前共排了0節課",
                                                 justify="center")

        self.formtext["M4"] = self.M4text
        self.formtext["M5"] = self.M5text
        self.formtext["M6"] = self.M6text
        self.formtext["M7"] = self.M7text
        self.formtext["M8"] = self.M8text
        self.formtext["M9"] = self.M9text
        self.formtext["M10"] = self.M10text
        self.formtext["MA"] = self.MAtext
        self.formtext["MB"] = self.MBtext
        self.formtext["MC"] = self.MCtext
        self.formtext["MD"] = self.MDtext
        self.formtext["T4"] = self.T4text
        self.formtext["T5"] = self.T5text
        self.formtext["T6"] = self.T6text
        self.formtext["T7"] = self.T7text
        self.formtext["T8"] = self.T8text
        self.formtext["T9"] = self.T9text
        self.formtext["T10"] = self.T10text
        self.formtext["TA"] = self.TAtext
        self.formtext["TB"] = self.TBtext
        self.formtext["TC"] = self.TCtext
        self.formtext["TD"] = self.TDtext
        self.formtext["W4"] = self.W4text
        self.formtext["W5"] = self.W5text
        self.formtext["W6"] = self.W6text
        self.formtext["W7"] = self.W7text
        self.formtext["W8"] = self.W8text
        self.formtext["W9"] = self.W9text
        self.formtext["W10"] = self.W10text
        self.formtext["WA"] = self.WAtext
        self.formtext["WB"] = self.WBtext
        self.formtext["WC"] = self.WCtext
        self.formtext["WD"] = self.WDtext
        self.formtext["R4"] = self.M4text
        self.formtext["R5"] = self.R5text
        self.formtext["R6"] = self.R6text
        self.formtext["R7"] = self.R7text
        self.formtext["R8"] = self.R8text
        self.formtext["R9"] = self.R9text
        self.formtext["R10"] = self.R10text
        self.formtext["RA"] = self.RAtext
        self.formtext["RB"] = self.RBtext
        self.formtext["RC"] = self.RCtext
        self.formtext["RD"] = self.RDtext
        self.formtext["F4"] = self.F4text
        self.formtext["F5"] = self.F5text
        self.formtext["F6"] = self.F6text
        self.formtext["F7"] = self.F7text
        self.formtext["F8"] = self.F8text
        self.formtext["F9"] = self.F9text
        self.formtext["F10"] = self.F10text
        self.formtext["FA"] = self.FAtext
        self.formtext["FB"] = self.FBtext
        self.formtext["FC"] = self.FCtext
        self.formtext["FD"] = self.FDtext

        # 以下為畫方格(非固定)
        self.M4 = self.mycanvas.create_rectangle(120, 65, 230, 130)
        self.M5 = self.mycanvas.create_rectangle(120, 130, 230, 195)
        self.M6 = self.mycanvas.create_rectangle(120, 195, 230, 260)
        self.M7 = self.mycanvas.create_rectangle(120, 260, 230, 325)
        self.M8 = self.mycanvas.create_rectangle(120, 325, 230, 390)
        self.M9 = self.mycanvas.create_rectangle(120, 390, 230, 455)
        self.M10 = self.mycanvas.create_rectangle(120, 455, 230, 520)
        self.MA = self.mycanvas.create_rectangle(120, 520, 230, 585)
        self.MB = self.mycanvas.create_rectangle(120, 585, 230, 650)
        self.MC = self.mycanvas.create_rectangle(120, 650, 230, 715)
        self.MD = self.mycanvas.create_rectangle(120, 715, 230, 780)
        self.T4 = self.mycanvas.create_rectangle(230, 65, 340, 130)
        self.T5 = self.mycanvas.create_rectangle(230, 130, 340, 195)
        self.T6 = self.mycanvas.create_rectangle(230, 195, 340, 260)
        self.T7 = self.mycanvas.create_rectangle(230, 260, 340, 325)
        self.T8 = self.mycanvas.create_rectangle(230, 325, 340, 390)
        self.T9 = self.mycanvas.create_rectangle(230, 390, 340, 455)
        self.T10 = self.mycanvas.create_rectangle(230, 455, 340, 520)
        self.TA = self.mycanvas.create_rectangle(230, 520, 340, 585)
        self.TB = self.mycanvas.create_rectangle(230, 585, 340, 650)
        self.TC = self.mycanvas.create_rectangle(230, 650, 340, 715)
        self.TD = self.mycanvas.create_rectangle(230, 715, 340, 780)
        self.W4 = self.mycanvas.create_rectangle(340, 65, 450, 130)
        self.W5 = self.mycanvas.create_rectangle(340, 130, 450, 195)
        self.W6 = self.mycanvas.create_rectangle(340, 195, 450, 260)
        self.W7 = self.mycanvas.create_rectangle(340, 260, 450, 325)
        self.W8 = self.mycanvas.create_rectangle(340, 325, 450, 390)
        self.W9 = self.mycanvas.create_rectangle(340, 390, 450, 455)
        self.W10 = self.mycanvas.create_rectangle(340, 455, 450, 520)
        self.WA = self.mycanvas.create_rectangle(340, 520, 450, 585)
        self.WB = self.mycanvas.create_rectangle(340, 585, 450, 650)
        self.WC = self.mycanvas.create_rectangle(340, 650, 450, 715)
        self.WD = self.mycanvas.create_rectangle(340, 715, 450, 780)
        self.R4 = self.mycanvas.create_rectangle(450, 65, 560, 130)
        self.R5 = self.mycanvas.create_rectangle(450, 130, 560, 195)
        self.R6 = self.mycanvas.create_rectangle(450, 195, 560, 260)
        self.R7 = self.mycanvas.create_rectangle(450, 260, 560, 325)
        self.R8 = self.mycanvas.create_rectangle(450, 325, 560, 390)
        self.R9 = self.mycanvas.create_rectangle(450, 390, 560, 455)
        self.R10 = self.mycanvas.create_rectangle(450, 455, 560, 520)
        self.RA = self.mycanvas.create_rectangle(450, 520, 560, 585)
        self.RB = self.mycanvas.create_rectangle(450, 585, 560, 650)
        self.RC = self.mycanvas.create_rectangle(450, 650, 560, 715)
        self.RD = self.mycanvas.create_rectangle(450, 715, 560, 780)
        self.F4 = self.mycanvas.create_rectangle(560, 65, 670, 130)
        self.F5 = self.mycanvas.create_rectangle(560, 130, 670, 195)
        self.F6 = self.mycanvas.create_rectangle(560, 195, 670, 260)
        self.F7 = self.mycanvas.create_rectangle(560, 260, 670, 325)
        self.F8 = self.mycanvas.create_rectangle(560, 325, 670, 390)
        self.F9 = self.mycanvas.create_rectangle(560, 390, 670, 455)
        self.F10 = self.mycanvas.create_rectangle(560, 455, 670, 520)
        self.FA = self.mycanvas.create_rectangle(560, 520, 670, 585)
        self.FB = self.mycanvas.create_rectangle(560, 585, 670, 650)
        self.FC = self.mycanvas.create_rectangle(560, 650, 670, 715)
        self.FD = self.mycanvas.create_rectangle(560, 715, 670, 780)

        self.the_seat["M4"] = self.M4
        self.the_seat["M5"] = self.M5
        self.the_seat["M6"] = self.M6
        self.the_seat["M7"] = self.M7
        self.the_seat["M8"] = self.M8
        self.the_seat["M9"] = self.M9
        self.the_seat["M10"] = self.M10
        self.the_seat["MA"] = self.MA
        self.the_seat["MB"] = self.MB
        self.the_seat["MC"] = self.MC
        self.the_seat["MD"] = self.MD
        self.the_seat["T4"] = self.T4
        self.the_seat["T5"] = self.T5
        self.the_seat["T6"] = self.T6
        self.the_seat["T7"] = self.T7
        self.the_seat["T8"] = self.T8
        self.the_seat["T9"] = self.T9
        self.the_seat["T10"] = self.T10
        self.the_seat["TA"] = self.TA
        self.the_seat["TB"] = self.TB
        self.the_seat["TC"] = self.TC
        self.the_seat["TD"] = self.TD
        self.the_seat["W4"] = self.W4
        self.the_seat["W5"] = self.W5
        self.the_seat["W6"] = self.W6
        self.the_seat["W7"] = self.W7
        self.the_seat["W8"] = self.W8
        self.the_seat["W9"] = self.W9
        self.the_seat["W10"] = self.W10
        self.the_seat["WA"] = self.WA
        self.the_seat["WB"] = self.WB
        self.the_seat["WC"] = self.WC
        self.the_seat["WD"] = self.WD
        self.the_seat["R4"] = self.R4
        self.the_seat["R5"] = self.R5
        self.the_seat["R6"] = self.R6
        self.the_seat["R7"] = self.R7
        self.the_seat["R8"] = self.R8
        self.the_seat["R9"] = self.R9
        self.the_seat["R10"] = self.R10
        self.the_seat["RA"] = self.RA
        self.the_seat["RB"] = self.RB
        self.the_seat["RC"] = self.RC
        self.the_seat["RD"] = self.RD
        self.the_seat["F4"] = self.F4
        self.the_seat["F5"] = self.F5
        self.the_seat["F6"] = self.F6
        self.the_seat["F7"] = self.F7
        self.the_seat["F8"] = self.F8
        self.the_seat["F9"] = self.F9
        self.the_seat["F10"] = self.F10
        self.the_seat["FA"] = self.FA
        self.the_seat["FB"] = self.FB
        self.the_seat["FC"] = self.FC
        self.the_seat["FD"] = self.FD

        # 以下為按鈕類控制
        self.Btnforadd = tk.Button(self.window, height=5, width=30, bg="Khaki", bd=2, text="新增人員資料",
                                   command=self.AddWindow)
        self.Btnforadd.place(x=1000, y=520)

        self.Btnfordel = tk.Button(self.window, height=5, bg="RosyBrown", width=30, bd=2, text="刪除人員資料",
                                   command=self.Delete)
        self.Btnfordel.place(x=720, y=520)

        self.Btnforset = tk.Button(self.window, height=5, bg="Coral", width=30, bd=2, text="設定人員資料",
                                   command=self.Setting)
        self.Btnforset.place(x=1000, y=650)

        self.Btnforexc = tk.Button(self.window, height=5, bg="OliveDrab", width=30, bd=2, text="匯出成excel檔",
                                   command=self.export_to_excel)
        self.Btnforexc.place(x=720, y=650)

        self.Btntolist = tk.Button(self.window, width=10, text="排入")
        self.Btntolist.configure(command=lambda: self.goform(self.entryforform.get()))
        self.Btntolist.place(x=1060, y=470)

        self.entryforform = tk.Entry(self.mycanvas, text="文字方塊", width=25)
        self.entryforform.place(x=950, y=482, anchor="center")

        self.window.mainloop()

    def AddWindow(self):
        self.isAttend = True
        # 哪一節能出席
        self.Attend = dict()

        self.seatBtn = dict()

        for j in range(5):
            day = ""
            if j == 0:
                day = "M"
            elif j == 1:
                day = "T"
            elif j == 2:
                day = "W"
            elif j == 3:
                day = "R"
            elif j == 4:
                day = "F"
            for i in range(4, 15):
                if i == 11:
                    self.Attend[day + "A"] = True
                elif i == 12:
                    self.Attend[day + "B"] = True
                elif i == 13:
                    self.Attend[day + "C"] = True
                elif i == 14:
                    self.Attend[day + "D"] = True
                else:
                    self.Attend[day + str(i)] = True

        self.newwin = tk.Toplevel(self.window)
        self.newwin.title("新增人員")
        self.newwin.geometry("960x680")
        self.newwin.resizable(width=False, height=False)

        self.can1 = tk.Canvas(self.newwin, width=442, height=660)  # 創造一個Canvas 442x660
        self.can1.pack(fill=tk.BOTH, expand=tk.YES)  # 排列

        self.can1.create_rectangle(10, 10, 442, 660)  # 畫外框

        # 姓名label
        self.labelname = tk.Label(self.can1, text="姓名：", width=30)
        self.labelname.place(x=590, y=155, anchor="center")

        # 姓名文字方塊
        self.entryname = tk.Entry(self.can1, text="這是文字方塊", width=15)
        self.entryname.place(x=680, y=155, anchor="center")

        # 多加課label
        self.labeladdclass = tk.Label(self.can1, text="選擇是否能排班可在左方點選\n或者打在下方 多個可用逗號分開 大小寫皆可\n(Ex. M5,r7,Fa)", width=50)
        self.labeladdclass.place(x=690, y=222, anchor="center")
        self.labelname1 = tk.Label(self.can1, text="[不可]排班的課:", width=30)
        self.labelname1.place(x=580, y=295, anchor="center")
        self.labelname2 = tk.Label(self.can1, text="[可]排班的課:", width=30)
        self.labelname2.place(x=580, y=355, anchor="center")

        # 多加課文字方塊
        self.entryaddclass = tk.Entry(self.can1, text="這", width=15)
        self.entryaddclass.place(x=680, y=295, anchor="center")

        self.entryaddclass2 = tk.Entry(self.can1, text="123", width=15)
        self.entryaddclass2.place(x=680, y=355, anchor="center")

        # 以下為畫橫線 第一橫條 為第一條以此類推 每條間隔55
        self.can1.create_line(10, 55, 442, 55)
        self.can1.create_line(10, 110, 442, 110)
        self.can1.create_line(10, 165, 442, 165)
        self.can1.create_line(10, 220, 442, 220)
        self.can1.create_line(10, 275, 442, 275)
        self.can1.create_line(10, 330, 442, 330)
        self.can1.create_line(10, 385, 442, 385)
        self.can1.create_line(10, 440, 442, 440)
        self.can1.create_line(10, 495, 442, 495)
        self.can1.create_line(10, 550, 442, 550)
        self.can1.create_line(10, 605, 442, 605)

        # 以下為畫直線 第一直條 為第一條以此類推 每條間隔72
        self.can1.create_line(82, 10, 82, 660)
        self.can1.create_line(154, 10, 154, 660)
        self.can1.create_line(226, 10, 226, 660)
        self.can1.create_line(298, 10, 298, 660)
        self.can1.create_line(370, 10, 370, 660)

        # 以下為畫第一格的斜線
        self.can1.create_line(10, 10, 82, 55)

        # 以下為橫行畫字 (固定)
        self.can1.create_text(118, 37, text='Monday\nM', font=('Arial', 10), justify="center")
        self.can1.create_text(190, 37, text='Tuesday\nT', font=('Arial', 10), justify="center")
        self.can1.create_text(262, 37, text='Wednesday\nW', font=('Arial', 10), justify="center")
        self.can1.create_text(334, 37, text='Thursday\nR', font=('Arial', 10), justify="center")
        self.can1.create_text(406, 37, text='Friday\nF', font=('Arial', 10), justify="center")

        # 以下為直行畫字 (固定)
        self.can1.create_text(30, 47, text='時間', font=('Arial', 9))
        self.can1.create_text(66, 26, text='星期', font=('Arial', 9))
        self.can1.create_text(46, 80, text='11:20~12:10\n第4節', font=('Arial', 9), justify="center")
        self.can1.create_text(46, 138, text='12:10~13:10\n第5節', font=('Arial', 9), justify="center")
        self.can1.create_text(46, 193, text='13:20~14:10\n第6節', font=('Arial', 9), justify="center")
        self.can1.create_text(46, 251, text='14:20~15:10\n第7節', font=('Arial', 9), justify="center")
        self.can1.create_text(46, 307, text='15:30~16:20\n第8節', font=('Arial', 9), justify="center")
        self.can1.create_text(46, 362, text='16:30~17:20\n第9節', font=('Arial', 9), justify="center")
        self.can1.create_text(46, 416, text='17:30~18:20\n第10節', font=('Arial', 9), justify="center")
        self.can1.create_text(46, 473, text='18:25~19:15\n第A節', font=('Arial', 9), justify="center")
        self.can1.create_text(46, 527, text='19:20~20:10\n第B節', font=('Arial', 9), justify="center")
        self.can1.create_text(46, 582, text='20:15~21:05\n第C節', font=('Arial', 9), justify="center")
        self.can1.create_text(46, 638, text='21:10~22:00\n第D節', font=('Arial', 9), justify="center")

        # 以下為每個按鈕 依照節數命名
        self.M4B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.M4B.configure(command=lambda: self.ChangeColor(self.M4B, "M4"))
        self.M4B.place(x=118, y=80, anchor="center")

        self.M5B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.M5B.configure(command=lambda: self.ChangeColor(self.M5B, "M5"))
        self.M5B.place(x=118, y=138, anchor="center")

        self.M6B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.M6B.configure(command=lambda: self.ChangeColor(self.M6B, "M6"))
        self.M6B.place(x=118, y=193, anchor="center")

        self.M7B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.M7B.configure(command=lambda: self.ChangeColor(self.M7B, "M7"))
        self.M7B.place(x=118, y=251, anchor="center")

        self.M8B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.M8B.configure(command=lambda: self.ChangeColor(self.M8B, "M8"))
        self.M8B.place(x=118, y=307, anchor="center")

        self.M9B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.M9B.configure(command=lambda: self.ChangeColor(self.M9B, "M9"))
        self.M9B.place(x=118, y=362, anchor="center")

        self.M10B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.M10B.configure(command=lambda: self.ChangeColor(self.M10B, "M10"))
        self.M10B.place(x=118, y=416, anchor="center")

        self.MAB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.MAB.configure(command=lambda: self.ChangeColor(self.MAB, "MA"))
        self.MAB.place(x=118, y=473, anchor="center")

        self.MBB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.MBB.configure(command=lambda: self.ChangeColor(self.MBB, "MB"))
        self.MBB.place(x=118, y=527, anchor="center")

        self.MCB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.MCB.configure(command=lambda: self.ChangeColor(self.MCB, "MC"))
        self.MCB.place(x=118, y=582, anchor="center")

        self.MDB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.MDB.configure(command=lambda: self.ChangeColor(self.MDB, "MD"))
        self.MDB.place(x=118, y=638, anchor="center")

        # Tuesday
        self.T4B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.T4B.configure(command=lambda: self.ChangeColor(self.T4B, "T4"))
        self.T4B.place(x=190, y=80, anchor="center")

        self.T5B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.T5B.configure(command=lambda: self.ChangeColor(self.T5B, "T5"))
        self.T5B.place(x=190, y=138, anchor="center")

        self.T6B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.T6B.configure(command=lambda: self.ChangeColor(self.T6B, "T6"))
        self.T6B.place(x=190, y=193, anchor="center")

        self.T7B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.T7B.configure(command=lambda: self.ChangeColor(self.T7B, "T7"))
        self.T7B.place(x=190, y=251, anchor="center")

        self.T8B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.T8B.configure(command=lambda: self.ChangeColor(self.T8B, "T8"))
        self.T8B.place(x=190, y=307, anchor="center")

        self.T9B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.T9B.configure(command=lambda: self.ChangeColor(self.T9B, "T9"))
        self.T9B.place(x=190, y=362, anchor="center")

        self.T10B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.T10B.configure(command=lambda: self.ChangeColor(self.T10B, "T10"))
        self.T10B.place(x=190, y=416, anchor="center")

        self.TAB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.TAB.configure(command=lambda: self.ChangeColor(self.TAB, "TA"))
        self.TAB.place(x=190, y=473, anchor="center")

        self.TBB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.TBB.configure(command=lambda: self.ChangeColor(self.TBB, "TB"))
        self.TBB.place(x=190, y=527, anchor="center")

        self.TCB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.TCB.configure(command=lambda: self.ChangeColor(self.TCB, "TC"))
        self.TCB.place(x=190, y=582, anchor="center")

        self.TDB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.TDB.configure(command=lambda: self.ChangeColor(self.TDB, "TD"))
        self.TDB.place(x=190, y=638, anchor="center")

        # Wednesday
        self.W4B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.W4B.configure(command=lambda: self.ChangeColor(self.W4B, "W4"))
        self.W4B.place(x=262, y=80, anchor="center")

        self.W5B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.W5B.configure(command=lambda: self.ChangeColor(self.W5B, "W5"))
        self.W5B.place(x=262, y=138, anchor="center")

        self.W6B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.W6B.configure(command=lambda: self.ChangeColor(self.W6B, "W6"))
        self.W6B.place(x=262, y=193, anchor="center")

        self.W7B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.W7B.configure(command=lambda: self.ChangeColor(self.W7B, "W7"))
        self.W7B.place(x=262, y=251, anchor="center")

        self.W8B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.W8B.configure(command=lambda: self.ChangeColor(self.W8B, "W8"))
        self.W8B.place(x=262, y=307, anchor="center")

        self.W9B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.W9B.configure(command=lambda: self.ChangeColor(self.W9B, "W9"))
        self.W9B.place(x=262, y=362, anchor="center")

        self.W10B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.W10B.configure(command=lambda: self.ChangeColor(self.W10B, "W10"))
        self.W10B.place(x=262, y=416, anchor="center")

        self.WAB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.WAB.configure(command=lambda: self.ChangeColor(self.WAB, "WA"))
        self.WAB.place(x=262, y=473, anchor="center")

        self.WBB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.WBB.configure(command=lambda: self.ChangeColor(self.WBB, "WB"))
        self.WBB.place(x=262, y=527, anchor="center")

        self.WCB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.WCB.configure(command=lambda: self.ChangeColor(self.WCB, "WC"))
        self.WCB.place(x=262, y=582, anchor="center")

        self.WDB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.WDB.configure(command=lambda: self.ChangeColor(self.WDB, "WD"))
        self.WDB.place(x=262, y=638, anchor="center")

        # Thursday
        self.R4B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.R4B.configure(command=lambda: self.ChangeColor(self.R4B, "R4"))
        self.R4B.place(x=334, y=80, anchor="center")

        self.R5B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.R5B.configure(command=lambda: self.ChangeColor(self.R5B, "R5"))
        self.R5B.place(x=334, y=138, anchor="center")

        self.R6B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.R6B.configure(command=lambda: self.ChangeColor(self.R6B, "R6"))
        self.R6B.place(x=334, y=193, anchor="center")

        self.R7B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.R7B.configure(command=lambda: self.ChangeColor(self.R7B, "R7"))
        self.R7B.place(x=334, y=251, anchor="center")

        self.R8B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.R8B.configure(command=lambda: self.ChangeColor(self.R8B, "R8"))
        self.R8B.place(x=334, y=307, anchor="center")

        self.R9B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.R9B.configure(command=lambda: self.ChangeColor(self.R9B, "R9"))
        self.R9B.place(x=334, y=362, anchor="center")

        self.R10B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.R10B.configure(command=lambda: self.ChangeColor(self.R10B, "R10"))
        self.R10B.place(x=334, y=416, anchor="center")

        self.RAB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.RAB.configure(command=lambda: self.ChangeColor(self.RAB, "RA"))
        self.RAB.place(x=334, y=473, anchor="center")

        self.RBB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.RBB.configure(command=lambda: self.ChangeColor(self.RBB, "RB"))
        self.RBB.place(x=334, y=527, anchor="center")

        self.RCB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.RCB.configure(command=lambda: self.ChangeColor(self.RCB, "RC"))
        self.RCB.place(x=334, y=582, anchor="center")

        self.RDB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.RDB.configure(command=lambda: self.ChangeColor(self.RDB, "RD"))
        self.RDB.place(x=334, y=638, anchor="center")

        # Friday
        self.F4B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.F4B.configure(command=lambda: self.ChangeColor(self.F4B, "F4"))
        self.F4B.place(x=406, y=80, anchor="center")

        self.F5B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.F5B.configure(command=lambda: self.ChangeColor(self.F5B, "F5"))
        self.F5B.place(x=406, y=138, anchor="center")

        self.F6B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.F6B.configure(command=lambda: self.ChangeColor(self.F6B, "F6"))
        self.F6B.place(x=406, y=193, anchor="center")

        self.F7B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.F7B.configure(command=lambda: self.ChangeColor(self.F7B, "F7"))
        self.F7B.place(x=406, y=251, anchor="center")

        self.F8B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.F8B.configure(command=lambda: self.ChangeColor(self.F8B, "F8"))
        self.F8B.place(x=406, y=307, anchor="center")

        self.F9B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.F9B.configure(command=lambda: self.ChangeColor(self.F9B, "F9"))
        self.F9B.place(x=406, y=362, anchor="center")

        self.F10B = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.F10B.configure(command=lambda: self.ChangeColor(self.F10B, "F10"))
        self.F10B.place(x=406, y=416, anchor="center")

        self.FAB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.FAB.configure(command=lambda: self.ChangeColor(self.FAB, "FA"))
        self.FAB.place(x=406, y=473, anchor="center")

        self.FBB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.FBB.configure(command=lambda: self.ChangeColor(self.FBB, "FB"))
        self.FBB.place(x=406, y=527, anchor="center")

        self.FCB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.FCB.configure(command=lambda: self.ChangeColor(self.FCB, "FC"))
        self.FCB.place(x=406, y=582, anchor="center")

        self.FDB = tk.Button(self.newwin, text="可排班", bg="lightgreen")
        self.FDB.configure(command=lambda: self.ChangeColor(self.FDB, "FD"))
        self.FDB.place(x=406, y=638, anchor="center")

        self.seatBtn["M4"] = self.M4B
        self.seatBtn["M5"] = self.M5B
        self.seatBtn["M6"] = self.M6B
        self.seatBtn["M7"] = self.M7B
        self.seatBtn["M8"] = self.M8B
        self.seatBtn["M9"] = self.M9B
        self.seatBtn["M10"] = self.M10B
        self.seatBtn["MA"] = self.MAB
        self.seatBtn["MB"] = self.MBB
        self.seatBtn["MC"] = self.MCB
        self.seatBtn["MD"] = self.MDB
        self.seatBtn["T4"] = self.T4B
        self.seatBtn["T5"] = self.T5B
        self.seatBtn["T6"] = self.T6B
        self.seatBtn["T7"] = self.T7B
        self.seatBtn["T8"] = self.T8B
        self.seatBtn["T9"] = self.T9B
        self.seatBtn["T10"] = self.T10B
        self.seatBtn["TA"] = self.TAB
        self.seatBtn["TB"] = self.TBB
        self.seatBtn["TC"] = self.TCB
        self.seatBtn["TD"] = self.TDB
        self.seatBtn["W4"] = self.W4B
        self.seatBtn["W5"] = self.W5B
        self.seatBtn["W6"] = self.W6B
        self.seatBtn["W7"] = self.W7B
        self.seatBtn["W8"] = self.W8B
        self.seatBtn["W9"] = self.W9B
        self.seatBtn["W10"] = self.W10B
        self.seatBtn["WA"] = self.WAB
        self.seatBtn["WB"] = self.WBB
        self.seatBtn["WC"] = self.WCB
        self.seatBtn["WD"] = self.WDB
        self.seatBtn["R4"] = self.R4B
        self.seatBtn["R5"] = self.R5B
        self.seatBtn["R6"] = self.R6B
        self.seatBtn["R7"] = self.R7B
        self.seatBtn["R8"] = self.R8B
        self.seatBtn["R9"] = self.R9B
        self.seatBtn["R10"] = self.R10B
        self.seatBtn["RA"] = self.RAB
        self.seatBtn["RB"] = self.RBB
        self.seatBtn["RC"] = self.RCB
        self.seatBtn["RD"] = self.RDB
        self.seatBtn["F4"] = self.F4B
        self.seatBtn["F5"] = self.F5B
        self.seatBtn["F6"] = self.F6B
        self.seatBtn["F7"] = self.F7B
        self.seatBtn["F8"] = self.F8B
        self.seatBtn["F9"] = self.F9B
        self.seatBtn["F10"] = self.F10B
        self.seatBtn["FA"] = self.FAB
        self.seatBtn["FB"] = self.FBB
        self.seatBtn["FC"] = self.FCB
        self.seatBtn["FD"] = self.FDB

        # 儲存方塊
        self.savebtn = tk.Button(self.newwin, text="儲存", width=40, height=5, bg="LightGrey")
        self.savebtn.configure(command=lambda: self.gosave(self.entryname.get(), self.Attend))
        self.savebtn.place(x=690, y=600, anchor="center")

        self.allcanbtn = tk.Button(self.newwin, text="全部可排班", width=12)
        self.allcanbtn.configure(command=self.allcanseat)
        self.allcanbtn.place(x=560, y=430)

        self.allcantbtn = tk.Button(self.newwin, text="全部不可排班", width=12)
        self.allcantbtn.configure(command=self.allcantseat)
        self.allcantbtn.place(x=730, y=430)

        self.cantchangebtn = tk.Button(self.newwin, text="排入", width=7)
        self.cantchangebtn.configure(command=lambda: self.cantchange(self.entryaddclass.get()))
        self.cantchangebtn.place(x=785, y=295, anchor="center")

        self.canchangebtn = tk.Button(self.newwin, text="排入", width=7)
        self.canchangebtn.configure(command=lambda: self.canchange(self.entryaddclass2.get()))
        self.canchangebtn.place(x=785, y=355, anchor="center")

        self.newwin.transient(self.window)
        self.newwin.grab_set()
        self.window.wait_window(self.newwin)

    def Setting(self):
        try:
            if self.nameforsetting == "":
                messagebox.showinfo("提示", "列表為空白或是沒有選擇人員!")
                return

            self.setwin = tk.Toplevel(self.window)
            self.setwin.title("設定人員")
            self.setwin.geometry("960x720")
            self.setwin.resizable(width=False, height=False)

            self.can2 = tk.Canvas(self.setwin, width=442, height=660)  # 創造一個Canvas 1024x768
            self.can2.pack(fill=tk.BOTH, expand=tk.YES)  # 排列

            self.can2.create_rectangle(10, 10, 442, 660)  # 畫外框

            self.SettingBtn = dict()

            setting_attend = self.WhosAttend[self.nameforsetting]

            # 多加課label
            self.labelsetname = tk.Label(self.can2, text="修改人員: " + self.nameforsetting, font=("Arial", 13), width=50)
            self.labelsetname.place(x=690, y=122, anchor="center")
            self.labelsetclass = tk.Label(self.can2, text="選擇是否能排班可在左方點選\n或者打在下方 多個可用逗號分開 大小寫皆可\n(Ex. M5,r7,Fa)",
                                          width=50)
            self.labelsetclass.place(x=690, y=222, anchor="center")
            self.labelsetname1 = tk.Label(self.can2, text="[不可]排班的課:", width=30)
            self.labelsetname1.place(x=580, y=295, anchor="center")
            self.labelsetname2 = tk.Label(self.can2, text="[可]排班的課:", width=30)
            self.labelsetname2.place(x=580, y=355, anchor="center")

            # 多加課文字方塊
            self.entrysetclass = tk.Entry(self.can2, text="這", width=15)
            self.entrysetclass.place(x=680, y=295, anchor="center")

            self.entrysetclass2 = tk.Entry(self.can2, text="123", width=15)
            self.entrysetclass2.place(x=680, y=355, anchor="center")

            # 以下為畫橫線 第一橫條 為第一條以此類推 每條間隔65
            self.can2.create_line(10, 75, 442, 75)
            self.can2.create_line(10, 140, 442, 140)
            self.can2.create_line(10, 205, 442, 205)
            self.can2.create_line(10, 270, 442, 270)
            self.can2.create_line(10, 335, 442, 335)
            self.can2.create_line(10, 400, 442, 400)
            self.can2.create_line(10, 465, 442, 465)
            self.can2.create_line(10, 530, 442, 530)
            self.can2.create_line(10, 595, 442, 595)

            # 以下為畫直線 第一直條 為第一條以此類推 每條間隔72
            self.can2.create_line(82, 10, 82, 660)
            self.can2.create_line(154, 10, 154, 660)
            self.can2.create_line(226, 10, 226, 660)
            self.can2.create_line(298, 10, 298, 660)
            self.can2.create_line(370, 10, 370, 660)

            # 以下為畫第一格的斜線
            self.can2.create_line(10, 10, 82, 75)

            # 以下為橫行畫字 (固定)
            self.can2.create_text(118, 43, text='Monday\nM', font=('Arial', 10), justify="center")
            self.can2.create_text(190, 43, text='Tuesday\nT', font=('Arial', 10), justify="center")
            self.can2.create_text(262, 43, text='Wednesday\nW', font=('Arial', 10), justify="center")
            self.can2.create_text(334, 43, text='Thursday\nR', font=('Arial', 10), justify="center")
            self.can2.create_text(406, 43, text='Friday\nF', font=('Arial', 10), justify="center")

            # 以下為直行畫字 (固定)
            self.can2.create_text(30, 57, text='時間', font=('Arial', 9))
            self.can2.create_text(66, 34, text='星期', font=('Arial', 9))
            self.can2.create_text(46, 108, text='12:10~13:10\n第5節', font=('Arial', 9), justify="center")
            self.can2.create_text(46, 173, text='13:20~14:10\n第6節', font=('Arial', 9), justify="center")
            self.can2.create_text(46, 238, text='14:20~15:10\n第7節', font=('Arial', 9), justify="center")
            self.can2.create_text(46, 303, text='15:30~16:20\n第8節', font=('Arial', 9), justify="center")
            self.can2.create_text(46, 368, text='16:30~17:20\n第9節', font=('Arial', 9), justify="center")
            self.can2.create_text(46, 433, text='17:30~18:20\n第10節', font=('Arial', 9), justify="center")
            self.can2.create_text(46, 498, text='18:25~19:15\n第A節', font=('Arial', 9), justify="center")
            self.can2.create_text(46, 563, text='19:20~20:10\n第B節', font=('Arial', 9), justify="center")
            self.can2.create_text(46, 628, text='20:15~21:05\n第C節', font=('Arial', 9), justify="center")

            # 以下為每個按鈕 依照節數命名
            self.M5S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.M5S.configure(command=lambda: self.ChangeColor(self.M5S, "M5"))
            self.M5S.place(x=118, y=108, anchor="center")

            self.M6S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.M6S.configure(command=lambda: self.ChangeColor(self.M6S, "M6"))
            self.M6S.place(x=118, y=173, anchor="center")

            self.M7S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.M7S.configure(command=lambda: self.ChangeColor(self.M7S, "M7"))
            self.M7S.place(x=118, y=238, anchor="center")

            self.M8S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.M8S.configure(command=lambda: self.ChangeColor(self.M8S, "M8"))
            self.M8S.place(x=118, y=303, anchor="center")

            self.M9S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.M9S.configure(command=lambda: self.ChangeColor(self.M9S, "M9"))
            self.M9S.place(x=118, y=368, anchor="center")

            self.M10S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.M10S.configure(command=lambda: self.ChangeColor(self.M10S, "M10"))
            self.M10S.place(x=118, y=433, anchor="center")

            self.MAS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.MAS.configure(command=lambda: self.ChangeColor(self.MAS, "MA"))
            self.MAS.place(x=118, y=498, anchor="center")

            self.MBS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.MBS.configure(command=lambda: self.ChangeColor(self.MBS, "MB"))
            self.MBS.place(x=118, y=563, anchor="center")

            self.MCS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.MCS.configure(command=lambda: self.ChangeColor(self.MCS, "MC"))
            self.MCS.place(x=118, y=628, anchor="center")

            # Tuesday
            self.T5S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.T5S.configure(command=lambda: self.ChangeColor(self.T5S, "T5"))
            self.T5S.place(x=190, y=108, anchor="center")

            self.T6S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.T6S.configure(command=lambda: self.ChangeColor(self.T6S, "T6"))
            self.T6S.place(x=190, y=173, anchor="center")

            self.T7S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.T7S.configure(command=lambda: self.ChangeColor(self.T7S, "T7"))
            self.T7S.place(x=190, y=238, anchor="center")

            self.T8S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.T8S.configure(command=lambda: self.ChangeColor(self.T8S, "T8"))
            self.T8S.place(x=190, y=303, anchor="center")

            self.T9S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.T9S.configure(command=lambda: self.ChangeColor(self.T9S, "T9"))
            self.T9S.place(x=190, y=368, anchor="center")

            self.T10S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.T10S.configure(command=lambda: self.ChangeColor(self.T10S, "T10"))
            self.T10S.place(x=190, y=433, anchor="center")

            self.TAS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.TAS.configure(command=lambda: self.ChangeColor(self.TAS, "TA"))
            self.TAS.place(x=190, y=498, anchor="center")

            self.TBS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.TBS.configure(command=lambda: self.ChangeColor(self.TBS, "TB"))
            self.TBS.place(x=190, y=563, anchor="center")

            self.TCS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.TCS.configure(command=lambda: self.ChangeColor(self.TCS, "TC"))
            self.TCS.place(x=190, y=628, anchor="center")

            # Wednesday 
            self.W5S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.W5S.configure(command=lambda: self.ChangeColor(self.W5S, "W5"))
            self.W5S.place(x=262, y=108, anchor="center")

            self.W6S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.W6S.configure(command=lambda: self.ChangeColor(self.W6S, "W6"))
            self.W6S.place(x=262, y=173, anchor="center")

            self.W7S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.W7S.configure(command=lambda: self.ChangeColor(self.W7S, "W7"))
            self.W7S.place(x=262, y=238, anchor="center")

            self.W8S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.W8S.configure(command=lambda: self.ChangeColor(self.W8S, "W8"))
            self.W8S.place(x=262, y=303, anchor="center")

            self.W9S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.W9S.configure(command=lambda: self.ChangeColor(self.W9S, "W9"))
            self.W9S.place(x=262, y=368, anchor="center")

            self.W10S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.W10S.configure(command=lambda: self.ChangeColor(self.W10S, "W10"))
            self.W10S.place(x=262, y=433, anchor="center")

            self.WAS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.WAS.configure(command=lambda: self.ChangeColor(self.WAS, "WA"))
            self.WAS.place(x=262, y=498, anchor="center")

            self.WBS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.WBS.configure(command=lambda: self.ChangeColor(self.WBS, "WB"))
            self.WBS.place(x=262, y=563, anchor="center")

            self.WCS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.WCS.configure(command=lambda: self.ChangeColor(self.WCS, "WC"))
            self.WCS.place(x=262, y=628, anchor="center")

            # Thursday 
            self.R5S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.R5S.configure(command=lambda: self.ChangeColor(self.R5S, "R5"))
            self.R5S.place(x=334, y=108, anchor="center")

            self.R6S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.R6S.configure(command=lambda: self.ChangeColor(self.R6S, "R6"))
            self.R6S.place(x=334, y=173, anchor="center")

            self.R7S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.R7S.configure(command=lambda: self.ChangeColor(self.R7S, "R7"))
            self.R7S.place(x=334, y=238, anchor="center")

            self.R8S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.R8S.configure(command=lambda: self.ChangeColor(self.R8S, "R8"))
            self.R8S.place(x=334, y=303, anchor="center")

            self.R9S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.R9S.configure(command=lambda: self.ChangeColor(self.R9S, "R9"))
            self.R9S.place(x=334, y=368, anchor="center")

            self.R10S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.R10S.configure(command=lambda: self.ChangeColor(self.R10S, "R10"))
            self.R10S.place(x=334, y=433, anchor="center")

            self.RAS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.RAS.configure(command=lambda: self.ChangeColor(self.RAS, "RA"))
            self.RAS.place(x=334, y=498, anchor="center")

            self.RBS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.RBS.configure(command=lambda: self.ChangeColor(self.RBS, "RB"))
            self.RBS.place(x=334, y=563, anchor="center")

            self.RCS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.RCS.configure(command=lambda: self.ChangeColor(self.RCS, "RC"))
            self.RCS.place(x=334, y=628, anchor="center")

            # Friday 
            self.F5S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.F5S.configure(command=lambda: self.ChangeColor(self.F5S, "F5"))
            self.F5S.place(x=406, y=108, anchor="center")

            self.F6S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.F6S.configure(command=lambda: self.ChangeColor(self.F6S, "F6"))
            self.F6S.place(x=406, y=173, anchor="center")

            self.F7S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.F7S.configure(command=lambda: self.ChangeColor(self.F7S, "F7"))
            self.F7S.place(x=406, y=238, anchor="center")

            self.F8S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.F8S.configure(command=lambda: self.ChangeColor(self.F8S, "F8"))
            self.F8S.place(x=406, y=303, anchor="center")

            self.F9S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.F9S.configure(command=lambda: self.ChangeColor(self.F9S, "F9"))
            self.F9S.place(x=406, y=368, anchor="center")

            self.F10S = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.F10S.configure(command=lambda: self.ChangeColor(self.F10S, "F10"))
            self.F10S.place(x=406, y=433, anchor="center")

            self.FAS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.FAS.configure(command=lambda: self.ChangeColor(self.FAS, "FA"))
            self.FAS.place(x=406, y=498, anchor="center")

            self.FBS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.FBS.configure(command=lambda: self.ChangeColor(self.FBS, "FB"))
            self.FBS.place(x=406, y=563, anchor="center")

            self.FCS = tk.Button(self.setwin, text="可排班", bg="lightgreen")
            self.FCS.configure(command=lambda: self.ChangeColor(self.FCS, "FC"))
            self.FCS.place(x=406, y=628, anchor="center")

            self.SettingBtn["M5"] = self.M5S
            self.SettingBtn["M6"] = self.M6S
            self.SettingBtn["M7"] = self.M7S
            self.SettingBtn["M8"] = self.M8S
            self.SettingBtn["M9"] = self.M9S
            self.SettingBtn["M10"] = self.M10S
            self.SettingBtn["MA"] = self.MAS
            self.SettingBtn["MB"] = self.MBS
            self.SettingBtn["MC"] = self.MCS
            self.SettingBtn["T5"] = self.T5S
            self.SettingBtn["T6"] = self.T6S
            self.SettingBtn["T7"] = self.T7S
            self.SettingBtn["T8"] = self.T8S
            self.SettingBtn["T9"] = self.T9S
            self.SettingBtn["T10"] = self.T10S
            self.SettingBtn["TA"] = self.TAS
            self.SettingBtn["TB"] = self.TBS
            self.SettingBtn["TC"] = self.TCS
            self.SettingBtn["W5"] = self.W5S
            self.SettingBtn["W6"] = self.W6S
            self.SettingBtn["W7"] = self.W7S
            self.SettingBtn["W8"] = self.W8S
            self.SettingBtn["W9"] = self.W9S
            self.SettingBtn["W10"] = self.W10S
            self.SettingBtn["WA"] = self.WAS
            self.SettingBtn["WB"] = self.WBS
            self.SettingBtn["WC"] = self.WCS
            self.SettingBtn["R5"] = self.R5S
            self.SettingBtn["R6"] = self.R6S
            self.SettingBtn["R7"] = self.R7S
            self.SettingBtn["R8"] = self.R8S
            self.SettingBtn["R9"] = self.R9S
            self.SettingBtn["R10"] = self.R10S
            self.SettingBtn["RA"] = self.RAS
            self.SettingBtn["RB"] = self.RBS
            self.SettingBtn["RC"] = self.RCS
            self.SettingBtn["F5"] = self.F5S
            self.SettingBtn["F6"] = self.F6S
            self.SettingBtn["F7"] = self.F7S
            self.SettingBtn["F8"] = self.F8S
            self.SettingBtn["F9"] = self.F9S
            self.SettingBtn["F10"] = self.F10S
            self.SettingBtn["FA"] = self.FAS
            self.SettingBtn["FB"] = self.FBS
            self.SettingBtn["FC"] = self.FCS

            for k, v in self.SettingBtn.items():
                if not (setting_attend.get(k)):
                    v["bg"] = "red"
                    v["text"] = "不可排班"

            self.allsetcanbtn = tk.Button(self.setwin, text="全部可排班", width=12)
            self.allsetcanbtn.configure(command=self.allsetcanseat)
            self.allsetcanbtn.place(x=560, y=430)

            self.allsetcantbtn = tk.Button(self.setwin, text="全部不可排班", width=12)
            self.allsetcantbtn.configure(command=self.allsetcantseat)
            self.allsetcantbtn.place(x=730, y=430)

            self.cantsetchangebtn = tk.Button(self.setwin, text="排入", width=7)
            self.cantsetchangebtn.configure(command=lambda: self.cantsetchange(self.entrysetclass.get()))
            self.cantsetchangebtn.place(x=785, y=295, anchor="center")

            self.cansetchangebtn = tk.Button(self.setwin, text="排入", width=7)
            self.cansetchangebtn.configure(command=lambda: self.cansetchange(self.entrysetclass2.get()))
            self.cansetchangebtn.place(x=785, y=355, anchor="center")

            self.setsavebtn = tk.Button(self.setwin, text="修改", width=40, height=5, bg="LightGrey")
            self.setsavebtn.configure(command=self.goset)
            self.setsavebtn.place(x=690, y=600, anchor="center")

            self.setwin.transient(self.window)
            self.setwin.grab_set()
            self.window.wait_window(self.setwin)
        except AttributeError:
            messagebox.showinfo("提示", "列表為空白或是沒有選擇人員!")
            return

    def validate(self):
        for k, v in self.formname.items():
            if self.nameforsetting in v and not (self.Attend.get(k)):
                Msg = messagebox.askyesno("提示", "目前修改不能出席的課裡已排入此人員\n是否要將其從此節課裡移除?")
                if Msg:
                    for j, i in self.formname.items():
                        if self.nameforsetting in i and not (self.Attend.get(j)):
                            temp1 = v.replace("\n" + self.name, "")
                            self.formname[j] = temp1
                            self.mycanvas.itemconfig(self.formtext[j], text=temp1)
                            tempnum = self.howmuch.get(self.nameforsetting)
                            self.howmuch[self.nameforsetting] = tempnum - 1
                    return True

                else:
                    return False
        return True

    def goset(self):
        # 檢查修改的課原本有沒有填入
        if self.validate():
            messagebox.showinfo("提示", "修改成功!")
            self.Flash()
            self.setwin.destroy()
        else:
            return

    def gosave(self, name, seat):
        if name == "":
            messagebox.showinfo("提示", "名子不能為空白!")
        elif name in self.WhosAttend:
            messagebox.showinfo("提示", "名子已經存在!")
        else:
            self.Saveinfo(name, seat)
            self.list1.insert(tk.END, name)
            self.howmuch[name] = 0
            messagebox.showinfo("提示", "儲存成功")
            self.entryname.delete(0, "end")
            self.newwin.destroy()

    def Saveinfo(self, name, seat):
        self.WhosAttend[name] = seat

    # 底下是設定人員視窗用的function
    def allsetcanseat(self):
        for i in self.SettingBtn.values():
            i["bg"] = "lightgreen"
            i["text"] = "可排班"
        for k in self.Attend.keys():
            self.Attend[k] = True

    def allsetcantseat(self):
        for i in self.SettingBtn.values():
            i["bg"] = "red"
            i["text"] = "不可排班"
        for k in self.Attend.keys():
            self.Attend[k] = False

    def cantsetchange(self, k):
        the_split = k.upper().split(',')
        for temp_class in the_split:
            # the_class 儲存節數 temp 為從formname取出之暫存檔做為加字
            # 檢查格子裡面的節數是否正確
            if not (temp_class in self.SettingBtn.keys()):
                messagebox.showinfo("提示", "輸入格式不符:[" + temp_class + "] 或是沒有這節課!")
                return
            elif the_split.count(temp_class) > 1:
                messagebox.showinfo("提示", "此節:[" + temp_class + "] 重複加入!")
                return
        for i in the_split:
            for k, v in self.SettingBtn.items():
                if i == k:
                    v["bg"] = "red"
                    v["text"] = "不可排班"
                    self.isAttend = False
                    self.Attend[i] = self.isAttend

    def cansetchange(self, k):
        the_split = k.upper().split(',')
        for temp_class in the_split:
            # the_class 儲存節數 temp 為從formname取出之暫存檔做為加字
            # 檢查格子裡面的節數是否正確
            if not (temp_class in self.SettingBtn.keys()):
                messagebox.showinfo("提示", "輸入格式不符:[" + temp_class + "] 或是沒有這節課!")
                return
            elif the_split.count(temp_class) > 1:
                messagebox.showinfo("提示", "此節:[" + temp_class + "] 重複加入!")
                return
        for i in the_split:
            for k, v in self.SettingBtn.items():
                if i == k:
                    v["bg"] = "lightgreen"
                    v["text"] = "可排班"
                    self.isAttend = True
                    self.Attend[i] = self.isAttend

    # 底下是新增人員視窗用的function
    def allcanseat(self):
        for i in self.seatBtn.values():
            i["bg"] = "lightgreen"
            i["text"] = "可排班"
        for k in self.Attend.keys():
            self.Attend[k] = True

    def allcantseat(self):
        for i in self.seatBtn.values():
            i["bg"] = "red"
            i["text"] = "不可排班"
        for k in self.Attend.keys():
            self.Attend[k] = False

    def cantchange(self, k):
        the_split = k.upper().split(',')
        for temp_class in the_split:
            # the_class 儲存節數 temp 為從formname取出之暫存檔做為加字
            # 檢查格子裡面的節數是否正確
            if not (temp_class in self.seatBtn.keys()):
                messagebox.showinfo("提示", "輸入格式不符:[" + temp_class + "] 或是沒有這節課!")
                return
            elif the_split.count(temp_class) > 1:
                messagebox.showinfo("提示", "此節:[" + temp_class + "] 重複加入!")
                return
        for i in the_split:
            for k, v in self.seatBtn.items():
                if i == k:
                    v["bg"] = "red"
                    v["text"] = "不可排班"
                    self.isAttend = False
                    self.Attend[i] = self.isAttend

    def canchange(self, k):
        the_split = k.upper().split(',')
        for temp_class in the_split:
            # the_class 儲存節數 temp 為從formname取出之暫存檔做為加字
            # 檢查格子裡面的節數是否正確
            if not (temp_class in self.seatBtn.keys()):
                messagebox.showinfo("提示", "輸入格式不符:[" + temp_class + "] 或是沒有這節課!")
                return
            elif the_split.count(temp_class) > 1:
                messagebox.showinfo("提示", "此節:[" + temp_class + "] 重複加入!")
                return
        for i in the_split:
            for k, v in self.seatBtn.items():
                if i == k:
                    v["bg"] = "lightgreen"
                    v["text"] = "可排班"
                    self.isAttend = True
                    self.Attend[i] = self.isAttend

    def ChangeColor(self, k, seat):
        if k["bg"] == "lightgreen":
            # 變成不可排班
            k["bg"] = "red"
            k["text"] = "不可排班"
            self.isAttend = False
            self.Attend[seat] = self.isAttend
        else:
            # 變成可排班
            k["bg"] = "lightgreen"
            k["text"] = "可排班"
            self.isAttend = True
            self.Attend[seat] = self.isAttend

    def listbox_service(self, event):
        try:
            self.Showseat(event)
            self.nameforsetting = self.list1.get(self.list1.curselection())
            # temp = event.widget
            # index = temp.curselection()
            # listname = temp.get(index)
            # self.Setting(listname)
        except tk.TclError:
            return

    def export_to_excel(self):
        try:
            # 定義名子
            fn = '吉他社排課.xlsx'

            the_alb = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]

            # 新增一個excel檔
            wb = openpyxl.Workbook()

            # 抓取目前的工作表
            ws = wb.active

            # 此為改變行高
            for i in range(1, 34):
                ws.row_dimensions[i].height = 25

            # 此為改變列寬
            for i in the_alb:
                ws.column_dimensions[i].width = 13

            # 合併儲存格
            ws.merge_cells('A1:B1')
            ws.merge_cells('A2:B2')
            ws.merge_cells('A3:B3')
            ws.merge_cells('A4:B4')
            ws.merge_cells('A5:B5')
            ws.merge_cells('A6:B6')
            ws.merge_cells('A7:B9')
            ws.merge_cells('C2:L6')
            ws.merge_cells('A10:B12')
            ws.merge_cells('A13:B15')
            ws.merge_cells('A16:B18')
            ws.merge_cells('A19:B21')
            ws.merge_cells('A22:B24')
            ws.merge_cells('A25:B27')
            ws.merge_cells('A28:B30')
            ws.merge_cells('A31:B33')
            ws.merge_cells('C1:D1')
            ws.merge_cells('C7:D9')
            ws.merge_cells('C10:D12')
            ws.merge_cells('C13:D15')
            ws.merge_cells('C16:D18')
            ws.merge_cells('C19:D21')
            ws.merge_cells('C22:D24')
            ws.merge_cells('C25:D27')
            ws.merge_cells('C28:D30')
            ws.merge_cells('C31:D33')
            ws.merge_cells('E1:F1')
            ws.merge_cells('E7:F9')
            ws.merge_cells('E10:F12')
            ws.merge_cells('E13:F15')
            ws.merge_cells('E16:F18')
            ws.merge_cells('E19:F21')
            ws.merge_cells('E22:F24')
            ws.merge_cells('E25:F27')
            ws.merge_cells('E28:F30')
            ws.merge_cells('E31:F33')
            ws.merge_cells('G1:H1')
            ws.merge_cells('G7:H9')
            ws.merge_cells('G10:H12')
            ws.merge_cells('G13:H15')
            ws.merge_cells('G16:H18')
            ws.merge_cells('G19:H21')
            ws.merge_cells('G22:H24')
            ws.merge_cells('G25:H27')
            ws.merge_cells('G28:H30')
            ws.merge_cells('G31:H33')
            ws.merge_cells('I1:J1')
            ws.merge_cells('I7:J9')
            ws.merge_cells('I10:J12')
            ws.merge_cells('I13:J15')
            ws.merge_cells('I16:J18')
            ws.merge_cells('I19:J21')
            ws.merge_cells('I22:J24')
            ws.merge_cells('I25:J27')
            ws.merge_cells('I28:J30')
            ws.merge_cells('I31:J33')
            ws.merge_cells('K1:L1')
            ws.merge_cells('K7:L9')
            ws.merge_cells('K10:L12')
            ws.merge_cells('K13:L15')
            ws.merge_cells('K16:L18')
            ws.merge_cells('K19:L21')
            ws.merge_cells('K22:L24')
            ws.merge_cells('K25:L27')
            ws.merge_cells('K28:L30')
            ws.merge_cells('K31:L33')

            # 修改儲存格資料 (固定)
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A1"].fill = PatternFill(fgColor="BEBEBE", fill_type="solid")
            ws['C1'].value = "一"
            ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["C1"].fill = PatternFill(fgColor="BEBEBE", fill_type="solid")
            ws['E1'].value = "二"
            ws["E1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["E1"].fill = PatternFill(fgColor="BEBEBE", fill_type="solid")
            ws['G1'].value = "三"
            ws["G1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["G1"].fill = PatternFill(fgColor="BEBEBE", fill_type="solid")
            ws['I1'].value = "四"
            ws["I1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["I1"].fill = PatternFill(fgColor="BEBEBE", fill_type="solid")
            ws['K1'].value = "五"
            ws["K1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["K1"].fill = PatternFill(fgColor="BEBEBE", fill_type="solid")

            ws['A2'].value = "時間"
            ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A2"].fill = PatternFill(fgColor="D3D3D3", fill_type="solid")
            ws['A3'].value = "1"
            ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A3"].fill = PatternFill(fgColor="D3D3D3", fill_type="solid")
            ws['A4'].value = "2"
            ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A4"].fill = PatternFill(fgColor="D3D3D3", fill_type="solid")
            ws['A5'].value = "3"
            ws["A5"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A5"].fill = PatternFill(fgColor="D3D3D3", fill_type="solid")
            ws['A6'].value = "4"
            ws["A6"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A6"].fill = PatternFill(fgColor="D3D3D3", fill_type="solid")
            ws['A7'].value = "12:20-13:10"
            ws["A7"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A7"].fill = PatternFill(fgColor="D3D3D3", fill_type="solid")
            ws['A10'].value = "13:20-14:10"
            ws["A10"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A10"].fill = PatternFill(fgColor="D3D3D3", fill_type="solid")
            ws['A13'].value = "14:20-15:10"
            ws["A13"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A13"].fill = PatternFill(fgColor="D3D3D3", fill_type="solid")
            ws['A16'].value = "15:30-16:20"
            ws["A16"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A16"].fill = PatternFill(fgColor="D3D3D3", fill_type="solid")
            ws['A19'].value = "16:30-17:20"
            ws["A19"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A19"].fill = PatternFill(fgColor="D3D3D3", fill_type="solid")
            ws['A22'].value = "17:30-18:20"
            ws["A22"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A22"].fill = PatternFill(fgColor="D3D3D3", fill_type="solid")
            ws['A25'].value = "18:25-19:15"
            ws["A25"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A25"].fill = PatternFill(fgColor="D3D3D3", fill_type="solid")
            ws['A28'].value = "19:20-20:10"
            ws["A28"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A28"].fill = PatternFill(fgColor="D3D3D3", fill_type="solid")
            ws['A31'].value = "20:15-21:05"
            ws["A31"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A31"].fill = PatternFill(fgColor="D3D3D3", fill_type="solid")

            # 修改儲存格資料 (非固定)
            ws['C7'].value = self.formname["M5"]
            ws["C7"].alignment = Alignment(horizontal="center", vertical="center")
            ws['C10'].value = self.formname["M6"]
            ws["C10"].alignment = Alignment(horizontal="center", vertical="center")
            ws['C13'].value = self.formname["M7"]
            ws["C13"].alignment = Alignment(horizontal="center", vertical="center")
            ws['C16'].value = self.formname["M8"]
            ws["C16"].alignment = Alignment(horizontal="center", vertical="center")
            ws['C19'].value = self.formname["M9"]
            ws["C19"].alignment = Alignment(horizontal="center", vertical="center")
            ws['C22'].value = self.formname["M10"]
            ws["C22"].alignment = Alignment(horizontal="center", vertical="center")
            ws['C25'].value = self.formname["MA"]
            ws["C25"].alignment = Alignment(horizontal="center", vertical="center")
            ws['C28'].value = self.formname["MB"]
            ws["C28"].alignment = Alignment(horizontal="center", vertical="center")
            ws['C31'].value = self.formname["MC"]
            ws["C31"].alignment = Alignment(horizontal="center", vertical="center")
            ws['E7'].value = self.formname["T5"]
            ws["E7"].alignment = Alignment(horizontal="center", vertical="center")
            ws['E10'].value = self.formname["T6"]
            ws["E10"].alignment = Alignment(horizontal="center", vertical="center")
            ws['E13'].value = self.formname["T7"]
            ws["E13"].alignment = Alignment(horizontal="center", vertical="center")
            ws['E16'].value = self.formname["T8"]
            ws["E16"].alignment = Alignment(horizontal="center", vertical="center")
            ws['E19'].value = self.formname["T9"]
            ws["E19"].alignment = Alignment(horizontal="center", vertical="center")
            ws['E22'].value = self.formname["T10"]
            ws["E22"].alignment = Alignment(horizontal="center", vertical="center")
            ws['E25'].value = self.formname["TA"]
            ws["E25"].alignment = Alignment(horizontal="center", vertical="center")
            ws['E28'].value = self.formname["TB"]
            ws["E28"].alignment = Alignment(horizontal="center", vertical="center")
            ws['E31'].value = self.formname["TC"]
            ws["E31"].alignment = Alignment(horizontal="center", vertical="center")
            ws['G7'].value = self.formname["W5"]
            ws["G7"].alignment = Alignment(horizontal="center", vertical="center")
            ws['G10'].value = self.formname["W6"]
            ws["G10"].alignment = Alignment(horizontal="center", vertical="center")
            ws['G13'].value = self.formname["W7"]
            ws["G13"].alignment = Alignment(horizontal="center", vertical="center")
            ws['G16'].value = self.formname["W8"]
            ws["G16"].alignment = Alignment(horizontal="center", vertical="center")
            ws['G19'].value = self.formname["W9"]
            ws["G19"].alignment = Alignment(horizontal="center", vertical="center")
            ws['G22'].value = self.formname["W10"]
            ws["G22"].alignment = Alignment(horizontal="center", vertical="center")
            ws['G25'].value = self.formname["WA"]
            ws["G25"].alignment = Alignment(horizontal="center", vertical="center")
            ws['G28'].value = self.formname["WB"]
            ws["G28"].alignment = Alignment(horizontal="center", vertical="center")
            ws['G31'].value = self.formname["WC"]
            ws["G31"].alignment = Alignment(horizontal="center", vertical="center")
            ws['I7'].value = self.formname["R5"]
            ws["I7"].alignment = Alignment(horizontal="center", vertical="center")
            ws['I10'].value = self.formname["R6"]
            ws["I10"].alignment = Alignment(horizontal="center", vertical="center")
            ws['I13'].value = self.formname["R7"]
            ws["I13"].alignment = Alignment(horizontal="center", vertical="center")
            ws['I16'].value = self.formname["R8"]
            ws["I16"].alignment = Alignment(horizontal="center", vertical="center")
            ws['I19'].value = self.formname["R9"]
            ws["I19"].alignment = Alignment(horizontal="center", vertical="center")
            ws['I22'].value = self.formname["R10"]
            ws["I22"].alignment = Alignment(horizontal="center", vertical="center")
            ws['I25'].value = self.formname["RA"]
            ws["I25"].alignment = Alignment(horizontal="center", vertical="center")
            ws['I28'].value = self.formname["RB"]
            ws["I28"].alignment = Alignment(horizontal="center", vertical="center")
            ws['I31'].value = self.formname["RC"]
            ws["I31"].alignment = Alignment(horizontal="center", vertical="center")
            ws['K7'].value = self.formname["F5"]
            ws["K7"].alignment = Alignment(horizontal="center", vertical="center")
            ws['K10'].value = self.formname["F6"]
            ws["K10"].alignment = Alignment(horizontal="center", vertical="center")
            ws['K13'].value = self.formname["F7"]
            ws["K13"].alignment = Alignment(horizontal="center", vertical="center")
            ws['K16'].value = self.formname["F8"]
            ws["K16"].alignment = Alignment(horizontal="center", vertical="center")
            ws['K19'].value = self.formname["F9"]
            ws["K19"].alignment = Alignment(horizontal="center", vertical="center")
            ws['K22'].value = self.formname["F10"]
            ws["K22"].alignment = Alignment(horizontal="center", vertical="center")
            ws['K25'].value = self.formname["FA"]
            ws["K25"].alignment = Alignment(horizontal="center", vertical="center")
            ws['K28'].value = self.formname["FB"]
            ws["K28"].alignment = Alignment(horizontal="center", vertical="center")
            ws['K31'].value = self.formname["FC"]
            ws["K31"].alignment = Alignment(horizontal="center", vertical="center")

            # 儲存
            wb.save(fn)
            messagebox.showinfo("提示", "匯出成功!")
        except Exception:
            messagebox.showerror("提示", "發生錯誤,匯出失敗!")

    def Showseat(self, event):
        try:
            temp = event.widget
            self.index = temp.curselection()
            self.name = temp.get(self.index)

            # 在主視窗顯示哪裡有空堂 用顏色區分
            # 先把出席紀錄(attend)從whosattend拿出來放進暫存
            self.temp_attend = self.WhosAttend.get(self.name)

            # 先刷新目前的表格
            self.Flash()

            # 從暫存中把出席紀錄拿出來做判斷
            for data, can in self.temp_attend.items():
                if not (can):
                    self.mycanvas.itemconfig(self.the_seat[data], fill="red")

            # 顯示此人排了幾節課
            texttemp = "-" + self.name + "-" + "目前排了" + str(self.howmuch.get(self.name)) + "節課"
            self.mycanvas.itemconfig(self.textnum, text=texttemp)
        except tk.TclError:
            return
        except Exception:
            messagebox.showerror("提示", "發生異常錯誤!")
            return

    def Delete(self):
        try:
            self.nameforsetting = ""
            # 刷新表格
            self.Flash()
            # 刪除 出席資料
            self.WhosAttend.pop(self.name)
            # 刪除 list裡面的字
            self.list1.delete("active")
            # 刪除 選幾節課的資料
            self.howmuch.pop(self.name)
            # 刪除表格裡面的字
            for k, v in self.formname.items():
                if self.name in v:
                    temp1 = v.replace("\n" + self.name, "")
                    self.formname[k] = temp1
                    self.mycanvas.itemconfig(self.formtext[k], text=temp1)
            self.mycanvas.itemconfig(self.textnum, text="刪除完成!")
        except AttributeError:
            messagebox.showinfo("提示", "列表為空白或是沒有選擇人員!")
            return

    def goform(self, the_class):
        try:
            the_split = the_class.upper().split(',')
            for wq in the_split:
                print(wq)
            for temp_class in the_split:
                # the_class 儲存節數 temp 為從formname取出之暫存檔做為加字
                # 檢查格子裡面的節數是否正確
                if not (temp_class in self.the_seat.keys()):
                    messagebox.showinfo("提示", "輸入格式不符:[" + temp_class + "] 或是沒有這節課!")
                    return
                if not (self.temp_attend.get(temp_class)):
                    messagebox.showinfo("提示", "此節:[" + temp_class + "] 不能出席!")
                    return
                if self.name in self.formname.get(temp_class):
                    messagebox.showinfo("提示", "此節:[" + temp_class + "] 已經排過!")
                    return
                if the_split.count(temp_class) > 1:
                    messagebox.showinfo("提示", "此節:[" + temp_class + "] 重複加入!")
                    return
            # 取出表格字裡面的值做加字的動作
            for temp_class in the_split:
                for i in self.formname.keys():
                    if i == temp_class:
                        # 把表格的字取出來 加上目前的人名(self.name)後 儲存到暫存變數temp1裡
                        temp1 = self.formname.get(temp_class) + "\n" + self.name
                        # 把暫存變數的值再回放到 表格裡面的字 也就是刷新的步驟
                        self.formname[temp_class] = temp1
                        # i儲存節數 把表格的字做改變
                        self.mycanvas.itemconfig(self.formtext[i], text=temp1)
            # 這裡儲存這個人選了多少節
            self.howmuch[self.name] = self.howmuch.get(self.name) + len(the_split)
            self.mycanvas.itemconfig(self.textnum, text="排入成功!")
            # 刷新表格
            self.Flash()
            self.entryforform.delete(0, "end")
        except AttributeError:
            messagebox.showinfo("提示", "列表為空白或是沒有選擇人員!")
            return
        except Exception:
            messagebox.showerror("提示", "發生異常錯誤!")
            return

    def Flash(self):
        self.temp_attend = self.WhosAttend.get(self.name)
        for key in self.temp_attend.keys():
            self.mycanvas.itemconfig(self.the_seat[key], fill="")


app = Mainwindow()
